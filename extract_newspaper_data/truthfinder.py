import random,time,datetime,difflib,os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook



def parse_phone_numbers(soup):
    pers_crec = []
    for elem1 in soup.findAll('section', {'id': 'contact1'}):
        for elem2 in elem1.findAll('div', {'class': 'section-content phone-content'}):
            for elem3 in elem2.findAll('div', {'class': 'record'}):
                for elem4 in elem3.findAll('div', {'class': 'record-name'}):
                    pers_ph = elem4.text
                    pers_ph = pers_ph.strip()
                    try:
                        pers_ph = pers_ph.replace('View Phone Details', '')
                        pers_ph = pers_ph.replace('Phone Number', '')
                    except:
                        pass
                    pers_crec.append(pers_ph)
    final_list = []
    for e1 in pers_crec:
        final_list.append(e1.strip())
    print(' , '.join(final_list))
    return final_list


def parse_emails(soup):
    pers_erec = []
    for elem1 in soup.findAll('section', {'id': 'contact1'}):
        for elem5 in elem1.findAll('div', {'class': 'section-content email-content'}):
            for elem6 in elem5.findAll('div', {'class': 'section-table-row-item email-address'}):
                for elem7 in elem6.findAll('span'):
                    # print(elem7.text)
                    pers_em = elem7.text
                    pers_em = pers_em.strip()
                    pers_erec.append(pers_em)

    print('Email')
    final_list = []
    for e1 in pers_erec:
        final_list.append(e1.strip())
    print(' , '.join(final_list))
    return final_list

def is_el_visible(driver,xpath):
    try:
        driver.find_element_by_xpath(xpath)
        return True
    except:
        return False

def wait_for_el_to_be_visible(xpath):
    while True:
        try:
            myElem = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                xpath)))
            time.sleep(random.uniform(.03, 1.1))
            return True
        except:
            driver.refresh()

def write_record_csv(record,fname,cols):
    fout = open(fname,'a')
    fout.write(get_record_string(record,cols))
    fout.close()


def get_record_string(d,cols):
    vals = [d.get(col,"") for col in cols]
    string = '\t'.join(vals)+'\n'
    return string


def create_csv_file(fname,cols):
    fout = open(fname, "w")
    string = '\t'.join(cols) + '\n'
    fout.write(string)
    fout.close()

def scrape_site(addresses, filename, columns, cities, states):
    flag=0
    for a in range(0, len(addresses) - 1):
        record = {}
        addr = addresses[a]
        city = cities[a]
        state = states[a]
        record["Address"] = addr
        record["City"] = city
        record["State"] = state

        time.sleep(random.uniform(.04, 1.5))
        a = driver.find_element_by_xpath('//input[@name="street"]')
        if flag:
            a = driver.find_element_by_xpath('//input[@name="street" and ancestor::form[@tab="location"]]')
        a.clear()
        for letters in addr:
            time.sleep(random.uniform(0.4,1.3))
            a.send_keys(letters)
        time.sleep(random.uniform(.3,2.1))

        c = driver.find_element_by_xpath('//input[@name="city"]')
        if flag:
            c = driver.find_element_by_xpath('//input[@name="city" and ancestor::form[@tab="location"]]')
        c.clear()
        for letters in city:
            time.sleep(random.uniform(0.4,1.3))
            c.send_keys(letters)
        time.sleep(random.uniform(.3,2.1))

        s = Select(driver.find_element_by_xpath('//select[@name="state"]'))
        if flag:
            s = Select(driver.find_element_by_xpath('//select[@name="state" and ancestor::form[@tab="location"]]'))

        time.sleep(random.uniform(.3,2.1))
        s.select_by_value(state)

        if flag:
            driver.find_element_by_xpath('//div[contains(text(),"Search") and ancestor::form[@tab="location"]]').click()
        else:
            driver.find_element_by_xpath('//button[contains(text(),"Search")]').click()
        wait_for_el_to_be_visible('//div[@class="nav-item active" and @data-section-name="map"]')

        driver.find_element_by_xpath('//div[@data-section-name="properties"]').click()

        wait_for_el_to_be_visible('//div[@class="nav-item active" and @data-section-name="properties"]')

        owner = driver.find_element_by_xpath('//div[@class="property-current-owner"]/h2').text
        record["Owner"] = owner
        time.sleep(random.uniform(.03, 1.1))

        driver.find_element_by_xpath('//div[@data-section-name="residents"]').click()
        wait_for_el_to_be_visible('//div[@class="nav-item active" and @data-section-name="residents"]')

        d = driver.find_elements_by_xpath('//span[contains(.,"Possible Current")]/../..//div[@class="section-table-row-item residents-name"]')
        if len(d)>0:
            dResidents = [t.text for t in d]
            m = difflib.SequenceMatcher
            dResSim = [m(None,owner, R).ratio() for R in dResidents]
            i=1
            if owner in dResidents:
                i=dResidents.index(owner)
            else:
                i=max(enumerate(dResSim), key=lambda p: p[1])[0]
            record["Resident"] = dResidents[i]
            time.sleep(random.uniform(1, 2))

            driver.find_element_by_xpath('(//span[contains(.,"Possible Current")]/../..//a)[{0}]'.format(i+1)).click()
            time.sleep(random.uniform(1, 2))

            wait_for_el_to_be_visible('//div[@id="section-inline-header-personal"]')

            html = driver.page_source
            soup = BeautifulSoup(html)

            phone_numbers = parse_phone_numbers(soup)
            emails = parse_emails(soup)

            for i, phone_number in enumerate(phone_numbers):
                record["Phone{0}".format(i)] = phone_number
            for i, email in enumerate(emails):
                record["Email{0}".format(i)] = email
        else:
            record["Resident"] = "N/A"
        print(record)
        write_record_csv(record, filename, columns)
        driver.find_element_by_xpath('//div[@class="search-bar"]').click()
        driver.find_element_by_xpath('//li[@data-tab="location"]').click()
        time.sleep(random.uniform(5.0, 7.0))
        flag=1

if __name__ == '__main__':
    options = Options()
    options.add_argument("user-data-dir=ChromeSession")
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(1000)
    driver.get('https://www.truthfinder.com/login')
    time.sleep(2)
    if is_el_visible(driver,"//input[@id='email']"):
        wait_for_el_to_be_visible("//input[@id='email']")
        t = driver.find_element_by_id("email")
        t.clear()
        name = 'fake.mail0@yahoo.com'
        for letters in name:
            time.sleep(random.uniform(0.1, 1.3))
            t.send_keys(letters)
        p = driver.find_element_by_id("password")
        p.clear()
        passw = 'October2019'
        for letters in passw:
            time.sleep(random.uniform(0.1, 1.3))
            p.send_keys(letters)
        time.sleep(10)
        driver.find_element_by_xpath('//button[@id="login-button"]').click()
    datetime = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    directoryPath = "truth"
    os.chdir(directoryPath)
    directoryPath = os.getcwd()
    for folders, sub_folders, file in os.walk(directoryPath):
            for name in file:
                if name.endswith(".xlsx"):
                    cities = []
                    addresses = []
                    states = []
                    filename = os.path.join(folders, name)
                    print(filename)

                    outfile = os.path.basename(filename)[:-5]
                    outfile = outfile + '-UPDATE-' + str(datetime) + '.csv'

                    wb = load_workbook(filename, data_only=True)
                    # sheet = wb.get_sheet_by_name('Sheet1')
                    ws = wb.active
                    first_column = ws['A']
                    second_column = ws['B']
                    third_column = ws['C']

                    for x in range(len(first_column)):
                        addresses.append(first_column[x].value)

                    for x in range(len(second_column)):
                        cities.append(second_column[x].value)
                    for x in range(len(third_column)):
                        states.append(third_column[x].value)

                    aclean = [x for x in addresses if x != None]
                    cleanaddresses = [x for x in aclean if x != 'Address']

                    cclean = [x for x in cities if x != None]
                    cleancities = [x for x in cclean if x != 'City']

                    sclean = [x for x in states if x != None]
                    cleanstates = [x for x in sclean if x != 'State']

                    if not cleanaddresses:
                        print("List is empty")
                    else:
                        columns = ["Address","City","State","Owner","Resident","Phone0","Phone1","Phone2","Phone3","Phone4","Phone5","Phone6","Phone7","Phone8","Phone9","Phone10","Email0","Email1","Email2","Email3"]
                        create_csv_file(outfile,columns)
                        wait_for_el_to_be_visible("//input[@name='firstName']")
                        driver.find_element_by_xpath('//span[contains(text(), "Search By Address")]').click()
                        scrape_site(cleanaddresses, outfile,columns, cleancities, cleanstates)
